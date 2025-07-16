import streamlit as st
import pandas as pd
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Page configuration
st.set_page_config(page_title="HackerRank Monthly Report Combiner", layout="wide")

# Custom CSS for better styling
st.markdown("""
<style>
    .main-header {
        text-align: center;
        color: #2E86AB;
        font-size: 2.5rem;
        margin-bottom: 1rem;
    }
    .sub-header {
        text-align: center;
        color: #666;
        font-size: 1.1rem;
        margin-bottom: 2rem;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 0.375rem;
        padding: 1rem;
        margin: 1rem 0;
    }
    .error-box {
        background-color: #f8d7da;
        border: 1px solid #f5c6cb;
        border-radius: 0.375rem;
        padding: 1rem;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

# Header
st.markdown('<h1 class="main-header">📊 HackerRank Monthly Report Combiner</h1>', unsafe_allow_html=True)
st.markdown('<p class="sub-header">Upload your weekly HackerRank CSV or Excel reports to generate a combined monthly report based on 6-digit roll numbers.</p>', unsafe_allow_html=True)

# File uploader
uploaded_files = st.file_uploader(
    "📁 Upload CSV or Excel files (Weekly Reports)",
    type=["csv", "xlsx"],
    accept_multiple_files=True,
    help="Upload multiple files containing weekly HackerRank scores. Each file should have a 6-digit roll number column."
)

def read_file(file):
    """Read CSV or Excel file and return DataFrame"""
    try:
        if file.name.endswith(".csv"):
            # Try different encodings
            encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
            for encoding in encodings:
                try:
                    content = file.getvalue().decode(encoding)
                    return pd.read_csv(io.StringIO(content))
                except UnicodeDecodeError:
                    continue
            # If all encodings fail, try with errors='ignore'
            content = file.getvalue().decode('utf-8', errors='ignore')
            return pd.read_csv(io.StringIO(content))
        elif file.name.endswith(".xlsx"):
            return pd.read_excel(file)
    except Exception as e:
        st.error(f"❌ Error reading file {file.name}: {str(e)}")
        return None

def standardize_columns(df):
    """Standardize column names by removing extra spaces and converting to lowercase"""
    df.columns = df.columns.str.strip().str.lower()
    return df

def find_roll_number_column(df):
    """Find the roll number column in the DataFrame"""
    # First, check for columns with 'roll' in the name
    for col in df.columns:
        if 'roll' in col.lower():
            return col
    
    # Then, check for columns with 6-digit numbers
    for col in df.columns:
        try:
            # Convert to string and check if most values are 6-digit numbers
            str_series = df[col].astype(str).str.strip()
            # Remove any NaN values
            str_series = str_series.dropna()
            if len(str_series) == 0:
                continue
            # Check if at least 70% of values are 6-digit numbers
            six_digit_count = str_series.str.match(r'^\d{6}$').sum()
            if six_digit_count / len(str_series) >= 0.7:
                return col
        except:
            continue
    
    return None

def clean_roll_numbers(df, roll_col):
    """Clean and validate roll numbers"""
    # Convert to string and strip whitespace
    df[roll_col] = df[roll_col].astype(str).str.strip()
    
    # Remove rows with invalid roll numbers
    valid_pattern = r'^\d{6}$'
    df = df[df[roll_col].str.match(valid_pattern, na=False)]
    
    return df

def identify_score_columns(df, roll_col):
    """Identify valid score columns (numeric, not serial numbers)"""
    # Get all numeric columns
    numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
    
    # Remove roll number column if it's in numeric columns
    if roll_col in numeric_cols:
        numeric_cols.remove(roll_col)
    
    # Remove serial number columns
    serial_patterns = ['s.no', 'serial', 'sno', 'sr.no', 'sr no', 'slno', 'sl.no', 'sl no']
    valid_cols = []
    
    for col in numeric_cols:
        col_lower = col.lower()
        is_serial = any(pattern in col_lower for pattern in serial_patterns)
        
        # Additional check: if column values are sequential (1, 2, 3, ...), it's likely a serial number
        if not is_serial:
            try:
                values = df[col].dropna().astype(int)
                if len(values) > 1:
                    # Check if values are sequential starting from 1
                    expected_sequence = list(range(1, len(values) + 1))
                    if list(values.sort_values()) == expected_sequence:
                        is_serial = True
            except:
                pass
        
        if not is_serial:
            valid_cols.append(col)
    
    return valid_cols

def create_styled_excel(df, filename):
    """Create a styled Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Report"
    
    # Add data to worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Style the header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E86AB")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply header styling
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add borders and center alignment for all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# Main processing logic
if uploaded_files:
    st.markdown("---")
    st.markdown("### 📋 Processing Files...")
    
    # Progress bar
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    weekly_data = []
    roll_number_col = None
    file_info = []
    
    for idx, file in enumerate(uploaded_files):
        status_text.text(f"Processing file {idx + 1}/{len(uploaded_files)}: {file.name}")
        progress_bar.progress((idx + 1) / len(uploaded_files))
        
        # Read file
        df = read_file(file)
        if df is None:
            st.stop()
        
        # Store original column count
        original_cols = len(df.columns)
        original_rows = len(df)
        
        # Standardize columns
        df = standardize_columns(df)
        
        # Find roll number column
        current_roll_col = find_roll_number_column(df)
        if not current_roll_col:
            st.error(f"❌ Could not detect a 6-digit roll number column in **{file.name}**. Please ensure there's a column with 6-digit roll numbers or rename the column to include 'roll'.")
            st.stop()
        
        # Set roll number column for consistency
        if roll_number_col is None:
            roll_number_col = 'roll_number'
        
        # Rename roll number column to standard name
        df.rename(columns={current_roll_col: roll_number_col}, inplace=True)
        
        # Clean roll numbers
        df = clean_roll_numbers(df, roll_number_col)
        
        if len(df) == 0:
            st.error(f"❌ No valid 6-digit roll numbers found in **{file.name}**")
            st.stop()
        
        # Identify score columns
        score_cols = identify_score_columns(df, roll_number_col)
        
        if len(score_cols) == 0:
            st.error(f"❌ No valid numeric score columns found in **{file.name}**")
            st.stop()
        
        # Create week-specific dataframe
        week_df = df[[roll_number_col] + score_cols].copy()
        
        # Rename score columns to include week number
        renamed_cols = {}
        for col in score_cols:
            renamed_cols[col] = f"File {idx + 1} - {col.title()}"
        
        week_df.rename(columns=renamed_cols, inplace=True)
        
        # Store file information
        file_info.append({
            'file': file.name,
            'original_rows': original_rows,
            'valid_rows': len(week_df),
            'score_columns': len(score_cols),
            'column_names': list(renamed_cols.values())
        })
        
        weekly_data.append(week_df)
    
    # Display file processing summary
    st.markdown("### 📊 File Processing Summary")
    summary_df = pd.DataFrame([
        {
            'File Name': info['file'],
            'Original Rows': info['original_rows'],
            'Valid Rows': info['valid_rows'],
            'Score Columns': info['score_columns']
        }
        for info in file_info
    ])
    st.dataframe(summary_df, use_container_width=True)
    
    # Merge all weekly data
    status_text.text("Combining weekly reports...")
    merged_df = weekly_data[0]
    
    for df in weekly_data[1:]:
        merged_df = pd.merge(merged_df, df, on=roll_number_col, how='outer')
    
    # Fill missing values and sort
    merged_df = merged_df.fillna("-")
    merged_df = merged_df.sort_values(by=roll_number_col)
    
    # Rename roll number column to display name
    merged_df.rename(columns={roll_number_col: 'Roll Number'}, inplace=True)
    
    # Clear progress indicators
    progress_bar.empty()
    status_text.empty()
    
    # Display success message
    st.success("✅ Successfully combined all weekly reports!")
    
    # Display statistics
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Students", len(merged_df))
    with col2:
        st.metric("Total Files", len(uploaded_files))
    with col3:
        st.metric("Total Columns", len(merged_df.columns))
    
    # Display combined data
    st.markdown("### 📈 Combined Monthly Report")
    st.dataframe(merged_df, use_container_width=True)
    
    # Download button
    st.markdown("### 📥 Download Report")
    excel_data = create_styled_excel(merged_df, "HackerRank_Monthly_Report.xlsx")
    
    st.download_button(
        label="📥 Download Combined Monthly Report (Excel)",
        data=excel_data,
        file_name="HackerRank_Monthly_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Download the combined report as a styled Excel file"
    )
    
    # Additional options
    st.markdown("### ⚙️ Additional Options")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("📊 Show Detailed Statistics"):
            st.markdown("#### 📋 Detailed File Information")
            for info in file_info:
                with st.expander(f"📄 {info['file']}"):
                    st.write(f"**Original Rows:** {info['original_rows']}")
                    st.write(f"**Valid Rows:** {info['valid_rows']}")
                    st.write(f"**Score Columns:** {info['score_columns']}")
                    st.write("**Column Names:**")
                    for col in info['column_names']:
                        st.write(f"- {col}")
    
    with col2:
        # CSV download option
        csv_data = merged_df.to_csv(index=False)
        st.download_button(
            label="📄 Download as CSV",
            data=csv_data,
            file_name="HackerRank_Monthly_Report.csv",
            mime="text/csv",
            help="Download the combined report as a CSV file"
        )

else:
    # Instructions when no files are uploaded
    st.markdown("### 📋 Instructions")
    st.markdown("""
    1. **Upload Files**: Select multiple CSV or Excel files containing weekly HackerRank scores
    2. **Roll Number Column**: Ensure each file has a column with 6-digit roll numbers (can be named anything containing 'roll' or just contain 6-digit numbers)
    3. **Score Columns**: The app will automatically detect numeric score columns and exclude serial numbers
    4. **Download**: Get a combined monthly report with all weeks' data
    
    **File Format Requirements:**
    - Files should be in CSV or Excel format
    - Must contain a column with 6-digit roll numbers
    - Should have numeric score columns
    - Serial number columns will be automatically excluded
    """)
    
    st.markdown("### 🎯 Features")
    st.markdown("""
    - **Automatic Column Detection**: Detects roll numbers and score columns automatically
    - **Serial Number Exclusion**: Automatically excludes serial number columns
    - **Flexible File Support**: Works with CSV and Excel files
    - **Data Validation**: Validates roll numbers and cleans data
    - **Styled Output**: Generates professionally formatted Excel reports
    - **Error Handling**: Comprehensive error handling for various file formats
    """)

